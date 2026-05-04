"""Populate a copy of the stucco bid template with the AI's takeoff totals.

The template (data/real/bid_template.xlsx) drives every material and labor
cell from three inputs:
  C3 = total yards (walls + ceiling)
  C4 = total wall yards
  C5 = total ceiling yards

Given a takeoff result we copy the workbook and write those three cells.
Excel recalculates everything else when reopened.
"""

from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl


DEFAULT_TEMPLATE = Path(__file__).resolve().parent.parent.parent / "data" / "real" / "bid_template.xlsx"


def fill_bid_template(
    project_name: str,
    wall_yards: float,
    ceiling_yards: float = 0.0,
    template_path: Path | str = DEFAULT_TEMPLATE,
    out_path: Path | str | None = None,
    sheet_name: str = "3 Coat",
) -> Path:
    template_path = Path(template_path)
    if not template_path.exists():
        raise FileNotFoundError(f"bid template not found: {template_path}")
    out = Path(out_path) if out_path else template_path.with_name(
        f"bid_{project_name.replace(' ', '_')}.xlsx"
    )
    shutil.copyfile(template_path, out)
    wb = openpyxl.load_workbook(out)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    ws["B1"] = project_name
    ws["C4"] = float(wall_yards)
    ws["C5"] = float(ceiling_yards)
    ws["C3"] = float(wall_yards) + float(ceiling_yards)

    wb.save(out)
    return out
